import openpyxl
import os
from datetime import datetime
import re
from openpyxl.cell.cell import Cell, MergedCell

class ExcelHandler:
    def __init__(self, template_path):
        self.template_path = template_path

    def _clean_price(self, price_text):
        """Convert price text to a float number."""
        if not price_text: return 0.0
        text = str(price_text).strip()
        text = text.replace("Kč", "").replace("EUR", "").replace("€", "").replace(" ", "")
        text = text.replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return 0.0

    def _parse_date(self, date_text):
        """Find and parse a date from text."""
        if not date_text: return None
        text = str(date_text).strip()
        text = text.replace(",", ".")  # Fix common OCR comma-vs-dot error
        match = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{2,4})", text)
        if match:
            d, m, y = match.groups()
            if len(y) == 2: y = "20" + y 
            try:
                return datetime.strptime(f"{d}.{m}.{y}", "%d.%m.%Y")
            except ValueError: 
                pass
        return None

    def _get_writable_cell(self, ws, row, col):
        """
        Return the writable cell at (row, col).

        If the cell is a MergedCell, find the top-left master cell instead.
        """
        cell = ws.cell(row=row, column=col)
        
        # Normal cell — return directly
        if isinstance(cell, Cell):
            return cell
        
        # Merged cell — find the master
        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
        
        return cell

    def add_invoice_entry(self, output_path, data_dict):
        print("-" * 50)
        print(f"DEBUG: Zpracovávám data pro Excel...")
        
        # 1. Load template or existing output file
        file_to_load = self.template_path
        if os.path.exists(output_path):
            file_to_load = output_path
        elif not os.path.exists(self.template_path):
            print(f"(!) CHYBA: Nenalezena šablona: {self.template_path}")
            return False

        try:
            wb = openpyxl.load_workbook(file_to_load)
            
            # 2. Select worksheet
            sheet_name = "Příjmy a výdaje"
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            
            if ws is None:
                print("(!) CHYBA: List nebyl nalezen.")
                return False

            # 3. Find the first empty row
            # (check column 3 — Price)
            row = 79
            max_search = 5000
            counter = 0
            
            while counter < max_search:
                cell = self._get_writable_cell(ws, row, 3)
                if cell.value is None:
                    break
                row += 1
                counter += 1
            
            print(f"   (Zapisuji na řádek {row})")

            # 4. Write data (using _get_writable_cell)

            # -- VENDOR (Column 2 / B) --
            vendor_val = data_dict.get('vendor') or data_dict.get('vendor_text')
            if vendor_val:
                cell = self._get_writable_cell(ws, row, 2)
                cell.value = str(vendor_val) # type: ignore
            else:
                print("(!) VAROVÁNÍ: Klíč 'vendor' je prázdný!")

            # -- PRICE (Column 3 / C) --
            price_val = data_dict.get('price') or data_dict.get('price_text')
            if price_val:
                cell = self._get_writable_cell(ws, row, 3)
                cell.value = self._clean_price(price_val) # type: ignore
                cell.number_format = '#,##0.00 "Kč"'

            # -- DATE (Column 5 / E) --
            date_val = data_dict.get('date') or data_dict.get('date_text')
            if date_val:
                cell = self._get_writable_cell(ws, row, 5)
                val_parsed = self._parse_date(date_val)
                if val_parsed:
                    cell.value = val_parsed # type: ignore
                    cell.number_format = 'd.m.yyyy'
                else:
                    cell.value = str(date_val) # type: ignore

            # -- FILENAME (Column 6 / F) --
            if data_dict.get('filename'):
                cell = self._get_writable_cell(ws, row, 6)
                cell.value = data_dict['filename'] # type: ignore

            # 5. Save workbook
            wb.save(output_path)
            print(f"OK: Uloženo do '{output_path}'")
            return True

        except PermissionError:
            print(f"(!) CHYBA: Soubor '{output_path}' je otevřený v Excelu! Zavřete jej.")
            return False
        except Exception as e:
            print(f"(!) CHYBA: {e}")
            return False

